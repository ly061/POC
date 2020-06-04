"""
The market matrix excel utils
Get some matrix data from excel file. "MY20-Demo-Matrix-V6.5.xlsx"
"""

__author__ = "Michael.Tian"

import json
import openpyxl

__matrix_file_name__ = "proj_poc/data/Custom Showcase Market v1.xlsx"
__matrix_social_file_name__ = "proj_poc/data/Custom Showcase_Footer_Social_Links.xlsx"

def get_social_matrix():
    """
    Get all soical link matrix
    locale column B
    Facebook Url column C
    Instagram Url column D
    Twitter Url column E
    :return:
    """
    wb = openpyxl.load_workbook(__matrix_social_file_name__)
    sheet = wb["RYI"]
    res = {}
    for index in range(2, 33):
        locale = sheet["B{}".format(index)].value.split()[0]
        facebook = sheet["C{}".format(index)].value
        instagram = sheet["D{}".format(index)].value
        twitter = sheet["E{}".format(index)].value
        res[locale] = [link for link in [facebook, instagram, twitter] if link]

    wb.close()
    return res

def get_all_locale():
    """
    Get all locale from matrix for booking process
    :return:
    """
    wb = openpyxl.load_workbook(__matrix_file_name__)
    sheet = wb["MATRIX + HD.COM SOURCE"]

    res = []
    for index in range(4, 49):
        locale = sheet["B{}".format(index)].value
        res.append(locale)

    return res

def get_bike_matrix():
    """
    Get All bike info matrix
    Touring: H-L
    Cruiser: M-P
    Dark Custom: Q-T
    Performance: U-Y
    :return:
    """
    wb = openpyxl.load_workbook(__matrix_file_name__)
    sheet = wb["MATRIX + HD.COM SOURCE"]
    bike_category = {
        "touring": ["H", "I", "J", "K", "L"],
        "cruiser": ["M", "N", "O", "P"],
        "dark_custom": ["Q", "R", "S", "T"],
        "performance": ["U", "V", "W", "X", "Y"]
    }

    # bike_column = ["K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V"]
    bike_column = [col for c in [bike_category[key] for key in bike_category.keys()] for col in c]
    bike_code = {}
    for col in bike_column:
        code = sheet["{}2".format(col)].value
        bike_code[col] = code

    res = {}
    for index in range(4, 49):
        locale = sheet["B{}".format(index)].value
        res[locale] = {}
        for key in bike_category.keys():
            #res[locale][key] = []
            tmp = []
            column = bike_category[key]
            for col in column:
                cc = sheet["{}{}".format(col, index)].value
                if cc and cc.startswith("Y"):
                    #res[locale][key].append(bike_code[col])
                    tmp.append(bike_code[col])
            res[locale][key] = tmp

    wb.close()
    return res

if __name__ == "__main__":
    print(json.dumps(get_social_matrix()))